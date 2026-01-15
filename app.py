import io
import os
import time
import json
import random
import logging
import requests
import pandas as pd
import streamlit as st
from PIL import Image
from openpyxl.utils import get_column_letter
from openpyxl.drawing.image import Image as XLImage
from openpyxl.styles import Alignment

log = logging.getLogger('catalogue_images_ia')
logging.basicConfig(level=logging.INFO)

from openpyxl import load_workbook
from openpyxl.utils.cell import coordinate_from_string
from openpyxl.utils import column_index_from_string

DEFAULT_HEADERS = {
    "User-Agent": "Mozilla/5.0 (compatible; CatalogueImagesIA/1.0)",
    "Accept": "image/avif,image/webp,image/apng,image/*,*/*;q=0.8",
}

def fetch_image_bytes(url: str, timeout=(10, 35), retries: int = 3):
    """Télécharge une image en bytes de manière robuste (headers, redirects, retries).
    Lève une exception si échec ou si le contenu n'est pas une image."""
    last_err = None
    for attempt in range(retries):
        try:
            r = requests.get(url, timeout=timeout, headers=DEFAULT_HEADERS, allow_redirects=True)
            r.raise_for_status()
            ctype = (r.headers.get("Content-Type") or "").lower()
            # Certains serveurs ne renvoient pas de content-type, on tolère.
            if ctype and not ctype.startswith("image/"):
                raise ValueError(f"URL ne renvoie pas une image (Content-Type={ctype})")
            return r.content, ctype
        except Exception as e:
            last_err = e
            # backoff + jitter
            time.sleep(0.6 * (2 ** attempt) + random.random() * 0.2)
    raise last_err


def extract_original_photo_spans(
    excel_bytes: bytes,
    rows_to_skip_top: int,
    photo_column_name: str,
    df_columns,
    sheet_name: str | None = None,
):
    """Extrait les images d'une *colonne* du fichier Excel source et retourne :
    - spans: liste de tuples (start_df_idx, end_df_idx, img_bytes)
      où l'image doit être insérée uniquement sur start_df_idx puis la cellule est mergée jusqu'à end_df_idx.
    Limitations : nécessite que les images soient ancrées dans la feuille (cas Excel classique).
    """
    if photo_column_name not in df_columns:
        return []

    # Pandas: header = première ligne après skiprows
    header_row = int(rows_to_skip_top) + 1  # 1-based row number in Excel
    data_start_row = header_row + 1         # 1-based

    # colonne Excel (1-based) correspondant à photo_column_name
    col_pos_0 = list(df_columns).index(photo_column_name)
    photo_col_excel = col_pos_0 + 1

    wb = load_workbook(io.BytesIO(excel_bytes))
    ws = wb[sheet_name] if sheet_name and sheet_name in wb.sheetnames else wb.active

    # helper: find merged range containing a cell
    def merged_span_for_cell(r: int, c: int):
        # r,c are 1-based
        for mr in ws.merged_cells.ranges:
            if mr.min_row <= r <= mr.max_row and mr.min_col <= c <= mr.max_col:
                return mr.min_row, mr.max_row
        return r, r

    spans = []
    images = getattr(ws, "_images", [])
    for img in images:
        try:
            anchor = img.anchor
            # openpyxl image anchor usually has ._from with 0-based indexes
            frm = getattr(anchor, "_from", None) or getattr(anchor, "from_", None)
            if frm is None:
                continue
            img_col = int(getattr(frm, "col", -1)) + 1  # to 1-based
            img_row = int(getattr(frm, "row", -1)) + 1  # to 1-based
            if img_col != photo_col_excel:
                continue

            min_r, max_r = merged_span_for_cell(img_row, img_col)

            start_idx = min_r - data_start_row
            end_idx = max_r - data_start_row
            if end_idx < 0:
                continue
            if start_idx < 0:
                start_idx = 0

            # récupérer bytes de l'image
            img_bytes = None
            if hasattr(img, "_data") and callable(getattr(img, "_data")):
                img_bytes = img._data()
            else:
                # fallback: certains openpyxl versions ont .ref / .path
                img_bytes = None

            if not img_bytes:
                continue

            spans.append((start_idx, end_idx, img_bytes))
        except Exception:
            continue

    # dédupliquer/normaliser spans (garder le plus long si overlap)
    spans.sort(key=lambda x: (x[0], -(x[1]-x[0])))
    merged = []
    for s,e,b in spans:
        if not merged:
            merged.append([s,e,b])
            continue
        ps,pe,pb = merged[-1]
        # si même start, garder le plus long
        if s == ps:
            if e > pe:
                merged[-1] = [s,e,b]
            continue
        merged.append([s,e,b])
    return [(s,e,b) for s,e,b in merged]


def reorder_columns_for_export(df: pd.DataFrame, first_cols: list[str]) -> pd.DataFrame:
    first = [c for c in first_cols if c in df.columns]
    rest = [c for c in df.columns if c not in first]
    return df[first + rest]

# ==========================
# CONFIG GLOBALE
# ==========================

IMAGE_COLS = ["image_url_1", "image_url_2", "image_url_3"]
SERPAPI_ENDPOINT = "https://serpapi.com/search.json"
CACHE_PATH = "image_cache.json"

# Login / mot de passe via variables d'environnement
VALID_USERNAME = os.getenv("APP_USERNAME", "admin")
VALID_PASSWORD = os.getenv("APP_PASSWORD", "change_me")

# Clés SerpAPI pré-configurées (liste séparée par des virgules)
# Exemple: SERPAPI_KEYS="cle1,cle2,cle3"
SERPAPI_KEYS_RAW = os.getenv("SERPAPI_KEYS", "").strip()


# ==========================
# AUTH + CAPTCHA
# ==========================

def init_session_state():
    if "authenticated" not in st.session_state:
        st.session_state.authenticated = False
    if "captcha_question" not in st.session_state:
        regenerate_captcha()


def regenerate_captcha():
    a = random.randint(1, 9)
    b = random.randint(1, 9)
    st.session_state.captcha_question = f"{a} + {b}"
    st.session_state.captcha_expected = a + b


def login_form():
    st.title("🔐 Accès sécurisé au générateur de catalogue d'images")
    st.write("Merci de vous authentifier pour utiliser l’outil.")

    username = st.text_input("Identifiant")
    password = st.text_input("Mot de passe", type="password")
    st.write(f"Captcha : **{st.session_state.captcha_question} = ?**")
    captcha_input = st.text_input("Réponse au captcha")

    if st.button("Se connecter"):
        ok_creds = (username == VALID_USERNAME and password == VALID_PASSWORD)
        ok_captcha = (
            captcha_input.strip().isdigit()
            and int(captcha_input.strip()) == st.session_state.captcha_expected
        )

        if ok_creds and ok_captcha:
            st.session_state.authenticated = True
            st.success("Connexion réussie ✅")
        else:
            st.error("Identifiants ou captcha incorrects.")
            regenerate_captcha()


# ==========================
# CACHE GLOBAL (fichier JSON)
# ==========================

def load_cache(cache_path: str = CACHE_PATH):
    if os.path.exists(cache_path):
        try:
            with open(cache_path, "r", encoding="utf-8") as f:
                return json.load(f)
        except Exception:
            return {}
    return {}


def save_cache(cache: dict, cache_path: str = CACHE_PATH):
    try:
        with open(cache_path, "w", encoding="utf-8") as f:
            json.dump(cache, f, ensure_ascii=False, indent=2)
    except Exception:
        # on ignore les erreurs de sauvegarde pour ne pas bloquer l'app
        pass


# ==========================
# SERPAPI KEY SELECTION
# ==========================

def get_configured_serpapi_keys():
    """
    Retourne une liste de clés issues de SERPAPI_KEYS (env),
    ex: SERPAPI_KEYS="key1,key2,key3"
    """
    if not SERPAPI_KEYS_RAW:
        return []
    keys = [k.strip() for k in SERPAPI_KEYS_RAW.split(",") if k.strip()]
    return keys


def serpapi_key_selector():
    """
    UI pour sélectionner une clé SerpAPI :
    - soit parmi les clés configurées dans SERPAPI_KEYS
    - soit en saisissant une nouvelle clé manuellement
    """
    configured_keys = get_configured_serpapi_keys()

    serpapi_key = None

    if configured_keys:
        options = []
        for i, key in enumerate(configured_keys):
            last4 = key[-4:] if len(key) >= 4 else key
            options.append(f"Clé {i+1} (****{last4})")

        options.append("Saisir une autre clé...")

        choice = st.selectbox(
            "Choisir une clé SerpAPI parmi la configuration ou saisir une nouvelle clé :",
            options=options,
        )

        if choice == "Saisir une autre clé...":
            serpapi_key = st.text_input(
                "Nouvelle clé SerpAPI",
                type="password",
                help="Cette clé n'est pas stockée, elle est utilisée uniquement pour cette session.",
            )
        else:
            idx = options.index(choice)
            serpapi_key = configured_keys[idx]
            st.info(f"Clé SerpAPI sélectionnée : {choice}")
    else:
        serpapi_key = st.text_input(
            "Clé API SerpAPI",
            type="password",
            help="Aucune clé préconfigurée. Crée une clé sur https://serpapi.com/ puis colle-la ici.",
        )

    return serpapi_key


# ==========================
# RECHERCHE D’IMAGES
# ==========================

def build_query(row: pd.Series, search_columns):
    parts = []
    for col in search_columns:
        value = row.get(col)
        if pd.notna(value):
            text = str(value).strip()
            if text:
                parts.append(text)
    return " ".join(parts)


def search_images_serpapi(query: str, serpapi_key: str, max_results: int = 3, return_raw: bool = False):
    """Recherche d'images via SerpAPI.
    Si return_raw=True, retourne (urls, raw_json) pour faciliter le debug UI."""
    params = {
        "engine": "google_images",
        "q": query,
        "num": max_results,
        "api_key": serpapi_key,
    }

    try:
        log.info("SerpAPI request: q=%s num=%s key=%s****", query[:120], max_results, (serpapi_key or "")[:6])
    except Exception:
        pass

    response = requests.get(SERPAPI_ENDPOINT, params=params, timeout=30)
    response.raise_for_status()

    try:
        data = response.json()
    except Exception as e:
        raise ValueError(f"Réponse SerpAPI non-JSON: {e} (status={response.status_code})")

    results = data.get("images_results", []) or []
    urls = []
    for r in results:
        url = r.get("original") or r.get("thumbnail")
        if url:
            urls.append(url)
        if len(urls) >= max_results:
            break

    if return_raw:
        return urls, data
    return urls


def enrich_df_with_images(
    df: pd.DataFrame,
    serpapi_key: str,
    search_columns,
    max_images: int = 3,
    sleep_seconds: float = 0.2,
    test_mode: bool = False,
    max_rows_test: int = 50,
    use_global_cache: bool = True,
    use_unique_key_cache: bool = True,
    unique_key_columns=None,
    debug_mode: bool = False,
    debug_logs: list | None = None,
):
    """
    Enrichit un DataFrame avec des URLs d'images.

    - use_global_cache : utilise un cache texte->images (image_cache.json).
    - use_unique_key_cache : réutilise les images pour les mêmes clés uniques (REF, REF+EAN, etc.).
    - unique_key_columns : liste de colonnes composant la clé unique (facultatif).
    """
    if not search_columns:
        st.error("Aucune colonne sélectionnée pour la recherche d'images.")
        return df, [], []

    if unique_key_columns is None:
        unique_key_columns = []

    if debug_logs is None:
        debug_logs = []

    def _dbg(msg: str, payload=None):
        if not debug_mode:
            return
        entry = {"t": time.strftime("%H:%M:%S"), "msg": msg}
        if payload is not None:
            entry["payload"] = payload
        debug_logs.append(entry)


    # Colonnes de sortie
    for i, col in enumerate(IMAGE_COLS[:max_images]):
        if col not in df.columns:
            df[col] = None

    # Chargement du cache global
    global_cache = load_cache() if use_global_cache else {}
    # Cache clé unique : clé composite -> liste d'URLs
    unique_cache = {} if use_unique_key_cache and unique_key_columns else {}

    log_entries = []

    total_rows = len(df)
    rows_to_process = min(max_rows_test, total_rows) if test_mode else total_rows

    progress_bar = st.progress(0)
    status_text = st.empty()

    for idx, (row_index, row) in enumerate(df.iterrows(), start=1):
        if test_mode and idx > rows_to_process:
            break

        query = build_query(row, search_columns).strip()

        # Construire la clé unique (composite) si configurée
        unique_key = None
        if use_unique_key_cache and unique_key_columns:
            unique_key = tuple(str(row.get(col, "")).strip() for col in unique_key_columns)

        urls = []
        source = None  # "unique_key_cache" / "global_cache" / "serpapi" / None

        if not query:
            status_text.text(f"Ligne {row_index}: requête vide, ignorée.")
            _dbg("Empty query - skipped", {"row": row_index, "search_columns": search_columns})
        else:
            # 1) Cache par clé unique (super optimisé)
            if unique_key is not None and unique_key in unique_cache:
                urls = unique_cache[unique_key]
                source = "unique_key_cache"
                _dbg("Unique-key cache hit", {"row": row_index, "unique_key": list(unique_key) if isinstance(unique_key, tuple) else unique_key})

            else:
                # 2) Cache global texte -> URLs
                if use_global_cache and query in global_cache:
                    urls = global_cache[query]
                    source = "global_cache"
                    _dbg("Global cache hit", {"row": row_index, "query": query})
                else:
                    # 3) Appel SerpAPI
                    try:
                        if debug_mode:
                            urls, raw = search_images_serpapi(
                                query=query,
                                serpapi_key=serpapi_key,
                                max_results=max_images,
                                return_raw=True,
                            )
                            _dbg("SerpAPI response", {
                                "row": row_index,
                                "query": query,
                                "images_results": len(raw.get("images_results", []) or []),
                                "status": (raw.get("search_metadata", {}) or {}).get("status"),
                                "error": raw.get("error"),
                                "raw_keys": list(raw.keys())[:25],
                            })
                        else:
                            urls = search_images_serpapi(
                                query=query,
                                serpapi_key=serpapi_key,
                                max_results=max_images,
                            )
                        source = "serpapi"
                        if use_global_cache:
                            global_cache[query] = urls
                        if not test_mode:
                            time.sleep(sleep_seconds)
                    except Exception as e:
                        status_text.text(f"Ligne {row_index}: ERREUR ({e})")
                        _dbg("SerpAPI error", {"row": row_index, "query": query, "error": str(e)})
                        try:
                            log.warning("SerpAPI error row=%s query=%s err=%s", row_index, query[:120], e)
                        except Exception:
                            pass
                        urls = []

                # On alimente le cache clé unique, même si la liste est vide
                if unique_key is not None and use_unique_key_cache:
                    unique_cache[unique_key] = urls

            # Remplissage des colonnes image_url_X
            for i, url in enumerate(urls):
                df.at[row_index, IMAGE_COLS[i]] = url

            status_text.text(
                f"Ligne {row_index}: {len(urls)} image(s) "
                f"{'via ' + source if source else ''} pour '{query[:60]}...'"
            )

            log_entries.append(
                {
                    "row_index": row_index,
                    "query": query,
                    "source": source,
                    "unique_key": unique_key,
                    "num_images": len(urls),
                    "urls": urls,
                }
            )

        progress_bar.progress(idx / rows_to_process)

    status_text.text("Recherche d'images terminée ✅")

    # Sauvegarde du cache global
    if use_global_cache:
        save_cache(global_cache)

    return df, log_entries, debug_logs




# ==========================
# EXCEL + IMAGES
# ==========================
def create_excel_with_embedded_images(
    df: pd.DataFrame,
    image_cols,
    orig_photo_col: str | None = None,
    orig_photo_spans=None,
    jpg_quality: int = 80,
    max_width: int = 80,
    max_height: int = 80,
    unique_key_columns=None,
):
    """
    Crée un fichier Excel en mémoire avec :
    - les données du DataFrame (toutes les lignes),
    - les images téléchargées, converties en JPG,
      insérées dans les cellules correspondant aux colonnes image_cols,
      en conservant les proportions d'origine,
    - si unique_key_columns est défini :
        * les images sont insérées une seule fois par combinaison de clé
          (sur la première ligne du groupe),
        * les cellules des colonnes de la clé + colonnes photos sont mergées
          verticalement sur les lignes du groupe.
    - formattage : contenu centré (H/V) + largeur colonnes ajustée + hauteur
      augmentée pour les lignes avec image.
    """
    if unique_key_columns is None:
        unique_key_columns = []

    output = io.BytesIO()

    with pd.ExcelWriter(output, engine="openpyxl") as writer:
        # 1) Écrire les données brutes
        df.to_excel(writer, index=False, sheet_name="Produits")
        wb = writer.book
        ws = wb["Produits"]
        failed_images = []  # [(col, url, error)]

        rows_with_image = set()

        # 2) Gestion des groupes par clé unique (si clé définie)
        groups = []
        if unique_key_columns:
            current_key = None
            start_excel_row = None
            prev_excel_row = None

            for row_idx, row in df.iterrows():
                excel_row = row_idx + 2
                key = tuple(str(row.get(col, "")).strip() for col in unique_key_columns)

                if key != current_key:
                    if current_key is not None and start_excel_row is not None:
                        groups.append((current_key, start_excel_row, prev_excel_row))
                    current_key = key
                    start_excel_row = excel_row

                prev_excel_row = excel_row

            if current_key is not None and start_excel_row is not None:
                groups.append((current_key, start_excel_row, prev_excel_row))

        # 3) Insertion des images
        # 3.a) Photos d'origine (si demandées) : insertion depuis le fichier source
        if orig_photo_col and orig_photo_spans:
            try:
                # retrouver l'index de colonne (1-based) dans la feuille de sortie
                orig_col_idx = None
                for j, cell in enumerate(ws[1], start=1):
                    if str(cell.value).strip() == str(orig_photo_col).strip():
                        orig_col_idx = j
                        break

                if orig_col_idx:
                    for start_idx, end_idx, img_content in orig_photo_spans:
                        if start_idx < 0 or start_idx >= len(df):
                            continue
                        end_idx = min(int(end_idx), len(df) - 1)
                        out_start_row = 2 + int(start_idx)
                        out_end_row = 2 + int(end_idx)

                        try:
                            pil_img = Image.open(io.BytesIO(img_content))
                            if pil_img.mode != "RGB":
                                pil_img = pil_img.convert("RGB")

                            orig_w, orig_h = pil_img.size
                            scale_w = max_width / orig_w
                            scale_h = max_height / orig_h
                            scale = min(scale_w, scale_h, 1.0)

                            new_w = int(orig_w * scale)
                            new_h = int(orig_h * scale)

                            img_bytes = io.BytesIO()
                            pil_img.save(img_bytes, format="JPEG", quality=jpg_quality, optimize=True)
                            img_bytes.seek(0)

                            xl_img = XLImage(img_bytes)
                            xl_img.width = new_w
                            xl_img.height = new_h

                            cell_ref = f"{get_column_letter(orig_col_idx)}{out_start_row}"
                            ws.add_image(xl_img, cell_ref)

                            # merge vertical pour conserver le rendu "une photo pour plusieurs lignes"
                            if out_end_row > out_start_row:
                                ws.merge_cells(
                                    start_row=out_start_row,
                                    start_column=orig_col_idx,
                                    end_row=out_end_row,
                                    end_column=orig_col_idx,
                                )
                        except Exception:
                            continue
            except Exception:
                pass


        if unique_key_columns and groups:
            # Une fois par groupe, sur la première ligne du groupe
            for key, start_row, end_row in groups:
                # Index pandas correspondant à la première ligne du groupe
                df_idx = start_row - 2
                if df_idx < 0 or df_idx >= len(df):
                    continue
                row = df.iloc[df_idx]

                for img_col in image_cols:
                    if img_col not in df.columns:
                        continue

                    url = row.get(img_col)
                    if not isinstance(url, str) or not url.strip():
                        continue

                    try:
                        img_content, _ctype = fetch_image_bytes(url)
                        pil_img = Image.open(io.BytesIO(img_content))
                        if pil_img.mode != "RGB":
                            pil_img = pil_img.convert("RGB")

                        orig_w, orig_h = pil_img.size

                        # Respecter max_width / max_height sans agrandir
                        scale_w = max_width / orig_w
                        scale_h = max_height / orig_h
                        scale = min(scale_w, scale_h, 1.0)

                        new_w = int(orig_w * scale)
                        new_h = int(orig_h * scale)

                        img_bytes = io.BytesIO()
                        pil_img.save(
                            img_bytes,
                            format="JPEG",
                            quality=jpg_quality,
                            optimize=True,
                        )
                        img_bytes.seek(0)

                        xl_img = XLImage(img_bytes)
                        xl_img.width = new_w
                        xl_img.height = new_h

                        col_idx = df.columns.get_loc(img_col) + 1
                        col_letter = get_column_letter(col_idx)
                        cell_ref = f"{col_letter}{start_row}"  # une seule fois, sur la 1re ligne

                        ws.add_image(xl_img, cell_ref)
                        rows_with_image.add(start_row)

                    except Exception as e:
                        failed_images.append((img_col, url, str(e)))
                        continue
        else:
            # Pas de clé unique : on met les images ligne par ligne
            for row_idx, row in df.iterrows():
                excel_row = row_idx + 2

                for img_col in image_cols:
                    if img_col not in df.columns:
                        continue

                    url = row.get(img_col)
                    if not isinstance(url, str) or not url.strip():
                        continue

                    try:
                        img_content, _ctype = fetch_image_bytes(url)
                        pil_img = Image.open(io.BytesIO(img_content))
                        if pil_img.mode != "RGB":
                            pil_img = pil_img.convert("RGB")

                        orig_w, orig_h = pil_img.size

                        scale_w = max_width / orig_w
                        scale_h = max_height / orig_h
                        scale = min(scale_w, scale_h, 1.0)

                        new_w = int(orig_w * scale)
                        new_h = int(orig_h * scale)

                        img_bytes = io.BytesIO()
                        pil_img.save(
                            img_bytes,
                            format="JPEG",
                            quality=jpg_quality,
                            optimize=True,
                        )
                        img_bytes.seek(0)

                        xl_img = XLImage(img_bytes)
                        xl_img.width = new_w
                        xl_img.height = new_h

                        col_idx = df.columns.get_loc(img_col) + 1
                        col_letter = get_column_letter(col_idx)
                        cell_ref = f"{col_letter}{excel_row}"

                        ws.add_image(xl_img, cell_ref)
                        rows_with_image.add(excel_row)

                    except Exception as e:
                        failed_images.append((img_col, url, str(e)))
                        continue

        # 4) Merge des cellules pour les groupes de même clé (sur les colonnes clé + colonnes photos)
        if unique_key_columns and groups:
            cols_to_merge = list(unique_key_columns)
            cols_to_merge.extend(image_cols)  # on merge toutes les colonnes photo

            for key, start_row, end_row in groups:
                if end_row <= start_row:
                    continue  # une seule ligne, rien à merger

                for col_name in cols_to_merge:
                    if col_name not in df.columns:
                        continue
                    col_idx = df.columns.get_loc(col_name) + 1
                    col_letter = get_column_letter(col_idx)
                    cell_range = f"{col_letter}{start_row}:{col_letter}{end_row}"
                    ws.merge_cells(cell_range)

        # 5) Mise en forme : centrage H/V + wrap + largeur colonnes + hauteur lignes
        # Alignement pour toutes les cellules du tableau
        for row in ws.iter_rows(
            min_row=1, max_row=ws.max_row,
            min_col=1, max_col=ws.max_column
        ):
            for cell in row:
                cell.alignment = Alignment(
                    horizontal="center",
                    vertical="center",
                    wrap_text=True,
                )

        # Largeur des colonnes en fonction du contenu
        for col_idx in range(1, ws.max_column + 1):
            col_letter = get_column_letter(col_idx)
            max_length = 0
            for cell in ws[col_letter]:
                if cell.value is not None:
                    max_length = max(max_length, len(str(cell.value)))
            # limite pour éviter des colonnes gigantesques
            ws.column_dimensions[col_letter].width = min(max_length * 1.2 + 2, 50)

        # Hauteur des lignes : plus grande pour celles qui ont une image
        base_height = 18
        img_height = max(max_height, 60)  # au moins 60 pour bien voir la chaussure etc.
        for row_idx in range(1, ws.max_row + 1):
            if row_idx in rows_with_image:
                ws.row_dimensions[row_idx].height = img_height
            else:
                ws.row_dimensions[row_idx].height = base_height

    try:
        st.session_state['failed_images_export'] = failed_images
    except Exception:
        pass

    output.seek(0)
    return output

# ==========================
# INTERFACE STREAMLIT
# ==========================

def app_main():
    st.set_page_config(
        page_title="Catalogue destockage avec images",
        layout="wide"
    )

    st.title("🧠 Catalogue de destockage avec images (textile, chaussures, parfums, etc.)")

    st.header("1. Clé SerpAPI")
    serpapi_key = serpapi_key_selector()

    col1, col2 = st.columns(2)
    with col1:
        max_images = st.slider(
            "Nombre d'images par produit",
            min_value=1,
            max_value=len(IMAGE_COLS),
            value=3,
            step=1,
        )
    with col2:
        sleep_seconds = st.slider(
            "Pause entre chaque requête (secondes, hors mode test)",
            min_value=0.0,
            max_value=1.0,
            value=0.2,
            step=0.1,
            help="Permet de lisser les appels à l'API si tu as beaucoup de lignes.",
        )

    st.header("2. Mode test (pour ajuster les paramètres)")

    test_mode = st.checkbox(
        "Activer le mode test (ne traiter que les X premières lignes et afficher le log détaillé)",
        value=True,
    )

    debug_mode = st.checkbox(
        "🪲 Mode debug (logs détaillés dans l'UI)",
        value=False,
        help="Affiche les requêtes SerpAPI envoyées, le nombre de résultats et les erreurs éventuelles."
    )

    max_rows_test = st.number_input(
        "Nombre de lignes à traiter en mode test",
        min_value=1,
        value=50,
        step=1,
    )

    st.header("3. Upload du fichier Excel")

    # Nouveau : nombre de lignes à ignorer en haut du fichier
    rows_to_skip_top = st.number_input(
        "Nombre de lignes à ignorer en haut du fichier (avant la ligne d'en-têtes)",
        min_value=0,
        value=0,
        step=1,
        help=(
            "Exemple : si vos en-têtes commencent à la ligne 3 d'Excel, "
            "indiquez 2 pour ignorer les 2 premières lignes (titres, logos, etc.)."
        ),
    )

    uploaded_file = st.file_uploader(
        "Choisis ton fichier Excel (.xlsx)",
        type=["xlsx"],
    )

    if uploaded_file is None:
        st.info("➡️ Uploade un fichier Excel pour commencer.")
        return

    try:
        # On ignore les X premières lignes, puis la première ligne restante devient la ligne d'en-têtes
        excel_bytes = uploaded_file.getvalue()
        df = pd.read_excel(io.BytesIO(excel_bytes), skiprows=int(rows_to_skip_top))
    except Exception as e:
        st.error(f"Erreur de lecture du fichier Excel : {e}")
        return
    
    st.success("Fichier bien reçu ✅")

    st.header("3.b Photos d'origine (optionnel)")
    keep_original_photos = st.checkbox("Garder les photos d'origine (si le fichier en contient)", value=False)
    original_photo_col_name = None
    orig_photo_spans = None
    if keep_original_photos:
        original_photo_col_name = st.selectbox(
            "Nom de la colonne qui contient les photos d'origine",
            options=list(df.columns),
        )
        # Extraction des images depuis le fichier Excel source (colonne unique, avec lignes fusionnées possibles)
        try:
            orig_photo_spans = extract_original_photo_spans(
                excel_bytes=excel_bytes,
                rows_to_skip_top=int(rows_to_skip_top),
                photo_column_name=original_photo_col_name,
                df_columns=df.columns,
                sheet_name=None,
            )
            if not orig_photo_spans:
                st.warning("Aucune image détectée dans cette colonne (ou extraction non supportée).")
            else:
                st.info(f"{len(orig_photo_spans)} bloc(s) de photo d'origine détecté(s) ✅")
        except Exception as e:
            st.warning(f"Impossible d'extraire les photos d'origine : {e}")
            orig_photo_spans = None

    st.subheader("Aperçu des premières lignes")
    st.dataframe(df.head())

    columns = list(df.columns)

    st.header("4. Options de cache")

    use_global_cache = st.checkbox(
        "Activer le cache global (texte de requête -> images) [fichier image_cache.json]",
        value=True,
    )

    use_unique_key_cache = st.checkbox(
        "Activer le cache par clé unique (REF, REF+COLOR, etc.) "
        "et fusionner les lignes partageant cette clé dans la sortie",
        value=True,
    )

    unique_key_columns = []
    if use_unique_key_cache:
        unique_key_columns = st.multiselect(
            "Colonnes composant la clé unique (ex : REF, ou REF + Color) :",
            options=columns,
            help=(
                "Exemple : REF seul, ou [REF, COLOR]. "
                "Pour une même clé unique, l'image sera recherchée une seule fois, "
                "puis les lignes seront fusionnées dans la sortie (une ligne par clé unique)."
            ),
        )

    st.header("5. Colonnes à utiliser pour la requête (ordre de priorité)")

    candidate_keywords = [
        "ean", "gtin", "ref", "code", "sku",
        "nom", "name", "brand", "marque", "description", "taille", "couleur"
    ]
    default_cols = [
        col for col in columns
        if any(kw in col.lower() for kw in candidate_keywords)
    ][:4]

    search_columns = st.multiselect(
        "Colonnes utilisées pour construire la requête (dans l'ordre) :",
        options=columns,
        default=default_cols,
        help="L'ordre compte : ex. [marque, nom, REF, description].",
    )

    st.header("6. Paramètres d’insertion des images dans Excel")

    col3, col4, col5 = st.columns(3)
    with col3:
        jpg_quality = st.slider(
            "Qualité JPG (1 = très compressé, 95 = haute qualité)",
            min_value=30,
            max_value=95,
            value=80,
            step=1,
            help="Compromis entre taille de fichier et qualité visuelle."
        )

    with col4:
        max_width = st.slider(
            "Largeur max des images (px)",
            min_value=40,
            max_value=200,
            value=80,
            step=5,
        )

    with col5:
        max_height = st.slider(
            "Hauteur max des images (px)",
            min_value=40,
            max_value=200,
            value=80,
            step=5,
        )

    st.header("7. Lancer le traitement")

    if not serpapi_key:
        st.info("➡️ Renseigne ou sélectionne une clé SerpAPI pour activer le bouton.")
        return

    if not search_columns:
        st.info("➡️ Sélectionne au moins une colonne de recherche.")
        return

    if st.button("🔍 Enrichir avec des images et les insérer dans l'Excel"):
        with st.spinner("Traitement en cours..."):
            # 1) Enrichissement avec images (et caches)
            df_enriched, log_entries, debug_logs = enrich_df_with_images(
                df.copy(),
                serpapi_key=serpapi_key,
                search_columns=search_columns,
                max_images=max_images,
                sleep_seconds=sleep_seconds,
                test_mode=test_mode,
                max_rows_test=max_rows_test,
                use_global_cache=use_global_cache,
                use_unique_key_cache=use_unique_key_cache,
                unique_key_columns=unique_key_columns,
                debug_mode=debug_mode,
                debug_logs=[],
            )
            # 2) Pas de fusion de lignes : on garde toutes les lignes
            df_output = df_enriched.copy()

            # Optionnel : ajouter une colonne pour les photos d'origine (pour comparaison)
            orig_photo_col_out = None
            if 'keep_original_photos' in locals() and keep_original_photos and orig_photo_spans:
                orig_photo_col_out = 'PHOTO_ORIG'
                if orig_photo_col_out not in df_output.columns:
                    df_output.insert(0, orig_photo_col_out, '')

            # Mettre les colonnes photos en premier (origine puis images trouvées)
            first_cols = []
            if orig_photo_col_out:
                first_cols.append(orig_photo_col_out)
            first_cols.extend(IMAGE_COLS[:max_images])
            df_output = reorder_columns_for_export(df_output, first_cols)

            # 3) Génération de l'Excel avec images intégrées
            excel_buffer = create_excel_with_embedded_images(
                df_output,
                image_cols=IMAGE_COLS[:max_images],
                orig_photo_col=orig_photo_col_out,
                orig_photo_spans=orig_photo_spans,

                jpg_quality=jpg_quality,
                max_width=max_width,
                max_height=max_height,
                unique_key_columns=unique_key_columns if use_unique_key_cache else None,
            )
        st.subheader("Aperçu des données enrichies (après fusion éventuelle)")
        st.dataframe(df_output.head())

        if test_mode:
            st.subheader("🧪 Log du mode test (requêtes, source, clés uniques, images)")
            if log_entries:
                log_df = pd.DataFrame(log_entries)
                st.dataframe(log_df)
            else:
                st.write("Aucun log (peut-être pas de requêtes valides).")
        # Diagnostic : logs debug (SerpAPI, cache, erreurs)
        if debug_mode:
            with st.expander("🪲 Logs debug (SerpAPI & pipeline)", expanded=False):
                if debug_logs:
                    st.dataframe(pd.DataFrame(debug_logs))
                else:
                    st.info("Aucun log debug (aucun appel API / aucune entrée).")

        failed = st.session_state.get('failed_images_export', [])
        if failed:
            st.warning(f"{len(failed)} image(s) n'ont pas pu être téléchargées/insérées (timeout, anti-bot, format...).")
            with st.expander("Voir le détail des échecs"):
                st.dataframe(pd.DataFrame(failed, columns=["colonne", "url", "erreur"]))

        st.download_button(
            label="⬇️ Télécharger le fichier Excel avec images intégrées",
            data=excel_buffer,
            file_name="produits_avec_images_integrees.xlsx",
            mime="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        )


def main():
    init_session_state()

    if not st.session_state.authenticated:
        login_form()
    else:
        app_main()


if __name__ == "__main__":
    main()
